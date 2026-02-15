import telebot
from telebot import types
import random
import openpyxl
from datetime import datetime

TOKEN = "8427218470:AAF9_sdfcFOJQcq5n34tkpKcMhh8Lxd5JXc"
GROUP_ID = -1003852199617
ADMIN_ID = 1028958055

bot = telebot.TeleBot(TOKEN)

user_step = {}
user_data = {}

# ===== MAHSULOTLAR =====
products = {
    "💧 5L Suv": 6000,
    "💧 10L Suv": 8000,
    "💧 18.9L Suv": 15000
}

# ===== EXCEL =====
def save_to_excel(order):
    try:
        wb = openpyxl.load_workbook("buyurtmalar.xlsx")
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Ism", "Raqam", "Mahsulot", "Soni", "Narx", "Manzil", "Sana"])

    ws = wb.active
    ws.append(order)
    wb.save("buyurtmalar.xlsx")

# ===== START =====
@bot.message_handler(commands=['start'])
def start(message):
    chat_id = message.chat.id

    if chat_id in user_data:
        show_menu(chat_id)
        return

    user_data[chat_id] = {}
    user_step[chat_id] = "name"

    bot.send_message(
        chat_id,
        "💧 Assalamu aleykum!\n\n"
        "AKOWATER botiga xush kelibsiz.\n\n"
        "Ismingizni kiriting:"
    )

# ===== ISM =====
@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "name")
def get_name(message):
    user_data[message.chat.id]["name"] = message.text
    user_step[message.chat.id] = "phone"

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    btn = types.KeyboardButton("📱 Raqamni yuborish", request_contact=True)
    markup.add(btn)

    bot.send_message(
        message.chat.id,
        "📱 Telefon raqamingizni yuboring:",
        reply_markup=markup
    )

# ===== TELEFON (MAJBURIY) =====
@bot.message_handler(content_types=['contact'])
def get_phone(message):
    if user_step.get(message.chat.id) != "phone":
        return

    user_data[message.chat.id]["phone"] = message.contact.phone_number
    user_step[message.chat.id] = "product"

    bot.send_message(message.chat.id, "✅ Raqam qabul qilindi.", reply_markup=types.ReplyKeyboardRemove())
    show_menu(message.chat.id)

@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "phone")
def block_text_phone(message):
    bot.send_message(message.chat.id, "❗ Pastdagi tugma orqali raqam yuboring.")

# ===== MENU =====
def show_menu(chat_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for p in products:
        markup.add(p)

    bot.send_message(
        chat_id,
        "💧 Mahsulotni tanlang:",
        reply_markup=markup
    )

# ===== MAHSULOT TANLASH =====
@bot.message_handler(func=lambda m: m.text in products)
def select_product(message):
    user_data[message.chat.id]["product"] = message.text
    user_step[message.chat.id] = "quantity"

    bot.send_message(message.chat.id, "📦 Nechta dona kerak?")

# ===== SONI =====
@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "quantity")
def get_quantity(message):
    if not message.text.isdigit():
        bot.send_message(message.chat.id, "❗ Faqat son kiriting.")
        return

    quantity = int(message.text)
    product = user_data[message.chat.id]["product"]
    price = products[product] * quantity

    user_data[message.chat.id]["quantity"] = quantity
    user_data[message.chat.id]["price"] = price
    user_step[message.chat.id] = "address"

    bot.send_message(
        message.chat.id,
        f"💰 Hisob narxi: {price:,} so‘m\n\n📍 Manzilingizni kiriting:"
    )

# ===== MANZIL VA YUBORISH =====
@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "address")
def finish_order(message):
    address = message.text
    chat_id = message.chat.id

    user_data[chat_id]["address"] = address
    order_id = random.randint(10000, 99999)

    data = user_data[chat_id]

    text = f"""
🆕 <b>YANGI BUYURTMA #{order_id}</b>

┏━━━━━━━━━━━━━━━┓
💧 <b>AKOWATER</b>
┗━━━━━━━━━━━━━━━┛

👤 Ism: {data['name']}
📱 Tel: {data['phone']}

🛒 Mahsulot: {data['product']}
📦 Soni: {data['quantity']}
💰 Narxi: {data['price']:,} so‘m

📍 Manzil: {address}
⏰ {datetime.now().strftime("%d-%m-%Y %H:%M")}
"""

    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("✅ Qabul qilindi", callback_data=f"accept_{order_id}"))

    bot.send_message(GROUP_ID, text, parse_mode="HTML", reply_markup=markup)

    save_to_excel([
        order_id,
        data["name"],
        data["phone"],
        data["product"],
        data["quantity"],
        data["price"],
        address,
        datetime.now().strftime("%d-%m-%Y %H:%M")
    ])

    bot.send_message(chat_id, "✅ Buyurtmangiz qabul qilindi!\nTez orada bog‘lanamiz.")
    show_menu(chat_id)

# ===== ADMIN PANEL =====
@bot.message_handler(commands=['admin'])
def admin_panel(message):
    if message.from_user.id != ADMIN_ID:
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("📊 Statistika", "📁 Excel yuklash")

    bot.send_message(message.chat.id, "👑 Admin panel:", reply_markup=markup)

@bot.message_handler(func=lambda m: m.text == "📊 Statistika")
def stats(message):
    if message.from_user.id != ADMIN_ID:
        return

    try:
        wb = openpyxl.load_workbook("buyurtmalar.xlsx")
        ws = wb.active
        total = ws.max_row - 1
    except:
        total = 0

    bot.send_message(message.chat.id, f"📦 Jami buyurtmalar: {total}")

@bot.message_handler(func=lambda m: m.text == "📁 Excel yuklash")
def send_excel(message):
    if message.from_user.id != ADMIN_ID:
        return

    try:
        file = open("buyurtmalar.xlsx", "rb")
        bot.send_document(message.chat.id, file)
    except:
        bot.send_message(message.chat.id, "❗ Excel fayl topilmadi.")

# ===== QABUL QILINDI =====
@bot.callback_query_handler(func=lambda call: call.data.startswith("accept_"))
def accept_order(call):
    order_id = call.data.split("_")[1]

    new_text = call.message.text + "\n\n✅ <b>Status:</b> Qabul qilindi"

    bot.edit_message_text(
        new_text,
        call.message.chat.id,
        call.message.message_id,
        parse_mode="HTML"
    )

    bot.answer_callback_query(call.id, "Buyurtma tasdiqlandi ✅")
    bot.infinity_polling()
